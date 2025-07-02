import cloudscraper
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl
import os
import time

# Excel file setup
file_path = "stock_list.xlsx"

# Create example Excel file if it doesn't exist
if not os.path.exists(file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stocks"
    ws['A1'] = "Symbol"
    ws.append(["AAPL"])
    ws.append(["MSFT"])
    print("🆕 Created stock_list.xlsx with sample symbols.")
    wb.save(file_path)

# Load workbook
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Create scraper with Cloudflare bypass
scraper = cloudscraper.create_scraper()

# Base URL of the finance site (hidden behind generic naming)
BASE_URL = "https://www.example-finance.com"

def get_top_10_gains(symbol):
    try:
        # Search for stock symbol
        search_url = f"{BASE_URL}/search/?q={symbol}"
        res = scraper.get(search_url, headers={"User-Agent": "Mozilla/5.0"})
        soup = BeautifulSoup(res.text, "html.parser")

        # Get first result link
        link = soup.select_one("a.js-inner-all-results-quote-item.row")
        if not link:
            print(f"[❌] No result for {symbol} on finance site")
            return []

        href = link.get("href")
        if not href:
            return []

        # Access historical data page
        hist_url = f"{BASE_URL}{href}-historical-data"
        res = scraper.get(hist_url, headers={"User-Agent": "Mozilla/5.0"})
        soup = BeautifulSoup(res.text, "html.parser")
        rows = soup.select("table tbody tr")

        # Parse date and close price
        data = []
        for row in rows:
            cols = row.find_all("td")
            if len(cols) < 2:
                continue
            try:
                date = datetime.strptime(cols[0].text.strip(), "%b %d, %Y")
                close = float(cols[1].text.strip().replace(",", ""))
                data.append((date, close))
            except:
                continue

        if len(data) < 2:
            return []

        data.sort()
        base_price = data[0][1]

        gains = []
        for date, price in data:
            pct = round(((price - base_price) / base_price) * 100, 2)
            gains.append((date.strftime("%Y-%m-%d"), pct))

        top_10 = sorted(gains, key=lambda x: x[1], reverse=True)[:10]
        return top_10

    except Exception as e:
        print(f"[❌] Error processing {symbol}: {e}")
        return []

# Loop through each row in Excel
for row in range(2, ws.max_row + 1):
    symbol = ws.cell(row=row, column=1).value
    if not symbol:
        continue

    print(f"[🔍] Fetching top gainers for: {symbol}")
    results = get_top_10_gains(symbol)
    if not results:
        continue

    for i, (date_str, gain) in enumerate(results):
        col = 2 + i * 2  # B=2, D=4, ..., T=20
        ws.cell(row=row, column=col).value = date_str
        ws.cell(row=row, column=col + 1).value = f"{gain}%"

    time.sleep(1.5)

# Save workbook
wb.save(file_path)
print("✅ All data saved with Top 10 gain entries.")
